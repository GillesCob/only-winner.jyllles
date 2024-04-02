import os
import pandas as pd
from datetime import datetime
import openpyxl

from html2image import Html2Image
from openpyxl import load_workbook


#ATTENTION : BIEN LIRE LES INFOS SUIVANTES : 
#Va transformer la syntaxe Midjourney en Prompt présents dans l'excel
#Attention à bien mettre à jour le mois concerné !

scrapping_month = "Mars"
actual_day=31
#actual_day = int(datetime.now().day)
excel_sheet = str(actual_day)
occurences=0


# Chemin vers le fichier Excel
chemin_excel = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/EXCEL/DATAS.xlsx'
# Lire les données Excel en utilisant pandas
df = pd.read_excel(chemin_excel, sheet_name=excel_sheet)
# Chemin vers le dossier contenant les images
dossier_images = "/Users/gillescobigo/Downloads"
# Chemin vers le dossier contenant les images renommées
dossier_images_renamed = f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/IMAGES_MIDJOURNEY"


classeur_month = openpyxl.load_workbook(chemin_excel)

excel_sheet_informations = classeur_month[excel_sheet]
#Je créé une liste avec les prompts Midjourney
prompt_midjourney_in_excel = [cell.value for cell in excel_sheet_informations['O'] if cell.value is not None]
#Idem avec les noms des cartes
cards_name_in_excel = [cell.value for cell in excel_sheet_informations['M'] if cell.value is not None]

# Parcourir les fichiers dans le dossier d'images
for nom_fichier in os.listdir(dossier_images):
    chemin_fichier = os.path.join(dossier_images, nom_fichier)

    # Vérifier si le fichier est un fichier PNG
    if nom_fichier.lower().endswith(".png"):
        nom_fichier = nom_fichier[:60]
        if nom_fichier in prompt_midjourney_in_excel:
            index_prompt = prompt_midjourney_in_excel.index(nom_fichier)
            nouveau_nom = cards_name_in_excel[index_prompt]
            nouveau_nom = nouveau_nom + ".png"
            os.rename(chemin_fichier, os.path.join(dossier_images_renamed, nouveau_nom))
            
            prompt_midjourney_in_excel.pop(index_prompt)
            cards_name_in_excel.pop(index_prompt)
            print(f"Nom changé de {nom_fichier} en {nouveau_nom}")
            occurences +=1
        else :
            print(f"{nom_fichier} n'est pas dans les prompts Midjourney fournis dans l'excel")
            
print(f"{occurences} images renommées et prêtes pour la génération de cartes")

occurences = 0



hti = Html2Image(size=(1000,1500))

#Données de départ
scrapping_month = "Mars"
actual_day_str = str(actual_day)
Feuille_datas = actual_day_str

#Chemins vers mes 2 fichiers excel
excel_data_file_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/EXCEL/DATAS.xlsx'
excel_BDD_INITIALE_file_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/BDD INITIALE.xlsx'


#Chargement des fichiers Excel
excel_month_workbook = load_workbook(excel_data_file_path)
excel_BDD_INITIALE_workbook = load_workbook(excel_BDD_INITIALE_file_path)

#Identification des feuilles utilisées par la suite
excel_sheet = excel_month_workbook[Feuille_datas]
nom_comp_of_sport_sheet = excel_BDD_INITIALE_workbook["COMP OF SPORT"] #Sert pour la donnée en haut de carte

#Mise en place des listes contenant les données récoltées dans les feuilles Excel
comp_of_sport_BDD_card_list = []
PROMPT_list = []
PROMPT_list = []
NOM_NFT_list = []
COMP_OF_SPORT_list = []

#------------------Récupération des données dans l'Excel BDD INITIALE-----------------------------------#
#Je récupère toutes les valeurs présentes dans la colonne A de la feuille "COMP OF SPORT"
comp_of_sport_BDD_list = [cell.value for cell in nom_comp_of_sport_sheet['A'] if cell.value is not None]
        
#Je récupère toutes les valeurs présentes dans la colonne C (traduction de comp of sport pour le haut de la carte)
comp_of_sport_BDD_card_list = [cell.value for cell in nom_comp_of_sport_sheet['C'] if cell.value is not None]
        
#------------------Récupération des données dans l'Excel DU JOUR-----------------------------------#
COMP_list = [cell.value for cell in excel_sheet['F'] if cell.value is not None]

SPORT_list = [cell.value for cell in excel_sheet['E'] if cell.value is not None]

PROMPT_list = [cell.value for cell in excel_sheet['L'] if cell.value is not None]

NOM_NFT_list = [cell.value for cell in excel_sheet['M'] if cell.value is not None]



for index_competition, competition in enumerate(COMP_list[1:], start=1):
    sport = SPORT_list[index_competition]
    prompt = PROMPT_list[index_competition]
    nft_name = NOM_NFT_list[index_competition]
    competition_of_sport = (f"{competition} of {sport}")
    if competition_of_sport in comp_of_sport_BDD_list:
        second_index_competition = comp_of_sport_BDD_list.index(competition_of_sport)
        competition_of_sport = comp_of_sport_BDD_card_list[second_index_competition]

    #Chemin vers les images Midjourney
    image_midjourney_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/IMAGES_MIDJOURNEY/{nft_name}.png'
    
    
    # Modèle de HTML avec CSS
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Rectangles</title>
        
        
        <style>
            /* Ajoutez votre CSS ici */
            .container {{
                position: fixed;
                top: 0;
                left: 0;
                display: flex;
                justify-content: center;
                align-items: center;
                width: 100%;
            }}

            .black-rectangle {{
                width: 1000px;
                height: 1500px;
                background-color: black;
            }}

            .orange-rectangle {{
                position: absolute;
                width: 975px;
                height: 1475px;
                background: linear-gradient(to top right, #d8aa3f, #927021);
            }}

            .competition {{
                position: absolute;
                top: 37.5px;
                left: 50%;
                transform: translateX(-50%);
                color: rgb(0, 0, 0);
                font-size: 40px;
                white-space: nowrap;
            }}

            .année {{
                position: absolute;
                top: 100px;
                left: 50%;
                transform: translateX(-50%);
                color: rgb(0, 0, 0);
                font-size: 50px;
            }}

            .fond-prompt {{
                position: absolute;
                width: 897px;
                height: 500px;
                bottom: 50px;
                background: hsla(0, 0%, 71%, 0.7);
                border-radius: 20px;
                border: 1px solid black;
            }}

            .prompt {{
                position: absolute;
                top: 1092px;
                left: 50%;
                transform: translateX(-50%);
                width: 850px;
                height: 343px;
                color: black;
                font-size: 40px;
                text-align: center;
                display: flex;
                justify-content: center;
                align-items: center;
                line-height: 50px;
            }}

            .image {{
                position: absolute;
                top: 178px;
                left: 50%;
                transform: translate(-50%);
                width: 900px;
                height: 900px;
                max-width: 100%;
                max-height: 100%;
            }}
        </style>
        
        
    </head>
    <body>
        <div class="container">
            <div class="black-rectangle"></div>
            <div class="orange-rectangle"></div>
            <div class="competition">{competition_of_sport}</div>
            <div class="année">2024</div>
            <div class="fond-prompt"></div>
            <img src="{image_midjourney_path}" alt="Image" class="image">
            <div class="prompt">{prompt}</div>
        </div>
    </body>
    </html>
    """
    
    # Chemin vers le dossier de sauvegarde (fichier général + fichier du jour pour faciliter l'import des images sur WP)
    save_folder_paths = [f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/NFT_READY", f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/NFT_READY/IMPORT WP/{actual_day} {scrapping_month}"]
    occurences +=1
    for save_folder_path in save_folder_paths :
        # Définir le dossier de sortie
        hti.output_path = save_folder_path

        # Prendre une capture d'écran avec le HTML généré pour la compétition actuelle
        hti.screenshot(html_str=html_template, save_as=f'{nft_name}.png')
        
        
print(f"{occurences} cartes générées")