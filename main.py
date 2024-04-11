from bs4 import BeautifulSoup
import requests
import pandas as pd
import openpyxl

from datetime import datetime
import time
import random
import dictionary
import re

import os

import json


#Seule variable à changer, mois à scrapper en français
months_scrapped = ["Avril"]
#months_scrapped = ['Janvier', 'Février', 'Mars']

webhook_url = 'https://discord.com/api/webhooks/1220361004479676456/ERaQqkUNyJgoJhYxpPRBznTX6LpF0M4as3E7IMuOa5Qhj7LrOuJoTyJ0v6RxqqMW-7sh'



for month in months_scrapped :
    scrapping_month = month
    actual_day = int(datetime.now().day)
    #actual_day = 31
    verif_date_event = actual_day
    date_event = "March 1st"


    #------------------------------------------------------------------------------------------------------------------------------------#
    #-------------------------------------------------ELEMENTS EN ENTREE-------------------------------------------------#

    #----------------------FONCTIONS----------------------#
    def rename_prompt_for_midjourney(name_NFT): #Modification du prompt pour correspondre au titre automatique lors de l'enregistrement d'une image Midjourney
        prompt_midjourney = name_NFT
        prompt_midjourney = prompt_midjourney.replace(" ", "_")
        prompt_midjourney = prompt_midjourney.replace("-", "")
        prompt_midjourney = prompt_midjourney.replace(",", "")
        prompt_midjourney = prompt_midjourney.replace("'", "")
        prompt_midjourney = prompt_midjourney.replace("(", "")
        prompt_midjourney = prompt_midjourney.replace(")", "")
        prompt_midjourney = prompt_midjourney.replace("ö", "o")
        prompt_midjourney = prompt_midjourney.replace("ó", "o")
        prompt_midjourney = prompt_midjourney.replace("ø", "")
        prompt_midjourney = prompt_midjourney.replace("ü", "u")
        prompt_midjourney = prompt_midjourney.replace("é", "e")
        prompt_midjourney = prompt_midjourney.replace("è", "e")
        prompt_midjourney = prompt_midjourney.replace("í", "i")
        prompt_midjourney = prompt_midjourney.replace("ï", "i")
        prompt_midjourney = prompt_midjourney.replace("á", "a")
        prompt_midjourney = prompt_midjourney.replace("å", "a")
        prompt_midjourney = prompt_midjourney.replace("ä", "a")
        prompt_midjourney = prompt_midjourney.replace("ß", "")
        prompt_midjourney = prompt_midjourney.replace("ñ", "n")
        prompt_midjourney = prompt_midjourney.replace("+", "")
        prompt_midjourney = "jyllles_" + prompt_midjourney
        prompt_midjourney = prompt_midjourney[:60]
        new_winners_one_sheet['Prompt_Midjourney'] = prompt_midjourney
        
        
    def winner_event_date_concordance(winner,date_event, url_event): #je créé winner-date. S'il est déjà dans la liste, j'append pour le montrer à la fin, sinon je l'ajoute dans la liste
        winner_event_date_concordance = f"{winner} - {date_event}"
        if winner_event_date_concordance in winner_and_date_event_list:
            multiple_winnings_same_day_list.append(f"{winner_event_date_concordance} - {url_event}")
        winner_and_date_event_list.append(winner_event_date_concordance)


    def card_name_without_accent(name_NFT):
        name_NFT = name_NFT.replace("(", "")
        name_NFT = name_NFT.replace(")", "")
        name_NFT = name_NFT.replace(" ", "-")
        name_NFT = name_NFT.replace(",", "")
        name_NFT = name_NFT.replace("'", "")
        name_NFT = name_NFT.replace("ö", "o")
        name_NFT = name_NFT.replace("Ö", "O")
        name_NFT = name_NFT.replace("ø", "o")
        name_NFT = name_NFT.replace("ò", "o")
        name_NFT = name_NFT.replace("ó", "o")
        name_NFT = name_NFT.replace("ü", "u")
        name_NFT = name_NFT.replace("ú", "u")
        name_NFT = name_NFT.replace("é", "e")
        name_NFT = name_NFT.replace("É", "E")
        name_NFT = name_NFT.replace("è", "e")
        name_NFT = name_NFT.replace("ã", "a")
        name_NFT = name_NFT.replace("ä", "a")
        name_NFT = name_NFT.replace("á", "a")
        name_NFT = name_NFT.replace("å", "a")
        name_NFT = name_NFT.replace("ý", "y")
        name_NFT = name_NFT.replace("ß", "s")
        name_NFT = name_NFT.replace("í", "i")
        name_NFT = name_NFT.replace("ï", "i")
        name_NFT = name_NFT.replace("---", "-")
        name_NFT = name_NFT.replace("--", "-")
        return name_NFT

    def create_short_winner(winner): #Diminuer le nombre de caractères afin d'avoir une cohérence visuelle sur la page d'achat des cartes (4 gagnants fait que le titre est trop long et ça décale tout)
        LONGUEUR_TITRE_WORDPRESS = 18
        if len(winner) > LONGUEUR_TITRE_WORDPRESS :
            short_winner = winner.split()
            short_winner = short_winner[:4]
            short_winner = " ".join(short_winner)
            if len(short_winner) > LONGUEUR_TITRE_WORDPRESS :
                short_winner = short_winner[:LONGUEUR_TITRE_WORDPRESS] + "..."
            else:
                pass
        else:
            short_winner = winner
        return short_winner
    
    def twitter_datas (competition_of_sport_index, winner):
        competition_of_sport_arobase = competition_of_sport_arobase_list[competition_of_sport_index]
        competition_of_sport_hashtag = competition_of_sport_hashtag_list[competition_of_sport_index]
        
        resultats_twitter = {}
        
        if competition_of_sport_arobase == "-":
            no_competition_arobase_list.append(competition_of_sport_traduction_value)
            resultats_twitter['arobase'] = None
        elif competition_of_sport_arobase == "/":
            resultats_twitter['arobase'] = ""
        else:
            resultats_twitter['arobase'] = competition_of_sport_arobase
            
        if competition_of_sport_hashtag == "-":
            no_competition_hashtag_list.append(competition_of_sport_traduction_value)
            resultats_twitter['hashtag'] = None
        elif competition_of_sport_hashtag == "/":
            resultats_twitter['hashtag'] = ""
        else:
            resultats_twitter['hashtag'] = competition_of_sport_hashtag
            
        if winner in winner_twitter_list:
            winner_twitter_index = winner_twitter_list.index(winner)
            resultats_twitter['winner_account'] = winner_arobase_twitter_list[winner_twitter_index]
            if resultats_twitter['winner_account'] == "-" :
                winner_for_twitter_list = winner.split(",")
                winner_for_twitter_list = ["#" + winner.replace(" ", "") for winner in winner_for_twitter_list]
                resultats_twitter['winner_account'] = ", ".join(winner_for_twitter_list)
        else:
            resultats_twitter['winner_account'] = ""
            twitter_account_list.append(winner)

        return resultats_twitter  

    #----------------------VARIABLES INITIALES----------------------#

    actual_year = str(datetime.now().year)
    url = "https://www.les-sports.info/calendrier-sport-2024-p0-62024.html"
    midjourney_parameters = ",oil painting style, colorful lighting,--ar 1:1 --c 100 --s 965 --v 5.1"
    COMPETITION_COUNTER = 0
    EVENT_COUNTER = 0
    EVENT_SPECIFIC_COUNTER = 0 #Sert à mettre un numéro qui démarre à 1 pour les nouvelles victoires (nécessaire pour second_rename et l'export PPT)
    SEUIL_CORRESPONDANCE = 20

    nom_sport_sheet = "SPORT"
    nom_competition_sheet = "COMPETITION"
    nom_event_sheet = "EVENT"
    nom_ignore_event_sheet = "IGNORE EVENT"
    nom_country_sheet = "COUNTRY"
    nom_city_sheet = "CITY"
    nom_date_sheet = "DATE"
    nom_abreviation_sheet = "ABREVIATION"
    nom_comp_of_sport_sheet = "COMP OF SPORT"
    nom_month_sheet = "MONTH"
    nom_twitter_sheet = "TWITTER"

    dossier_NFT = f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/NFT_READY" #Dossier contenant tous les NFT déjà réalisés


    #----------------------LISTES----------------------#
    all_month_winners_list = [] #Liste contenant toutes les infos pour chaque event pour tout le mois
    winners_without_nft_list = [] #Contient tous les données concernant les vainqueurs ayant déjà leur carte finalisée

    winners_with_nft_list = [] #Contient tous les données concernant les vainqueurs ayant déjà leur carte finalisée

    data_for_wordpress_list = [] #Liste de toutes les infos pour l'import Product sur WP

    #----------------------LISTES LIEES A DE LA VERIFICATIONS----------------------#
    winner_and_date_event_list = [] #j'ajoute dans cette liste les athlètes ayant gagné plusieurs fois le même jour

    events_ok_list = [] #Liste des events passés par les différents tamis
    all_events_in_page_list = [] #Je mets les events de nomargin et centre



    #----------------------LISTES POUR LE PRINT FINAL----------------------#
    no_country_list = []
    no_city_list = []
    no_sport_list = []
    no_sport_competition_list = []
    no_competition_of_sport_list = []
    no_event_list = []
    just_men_woman_list = []
    no_event_probably_empty_list = []
    no_date_event_list = []
    no_winner_list = []
    no_abr_list = []
    multiple_winnings_same_day_list = []
    cards_ignored_list = []
    mixed_event_list = []
    
    #Listes concernant les @ et # Twitter
    twitter_account_list = []
    no_competition_arobase_list = []
    no_competition_hashtag_list = []
    
    #Liste pour Discord
    discord_prompt_list = []

    #----------------------VERIFS EXCEL----------------------#
    #1 - Est-ce que le fichier Excel existe ?
    nom_excel_month = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/EXCEL/DATAS.xlsx'
    nom_exel_bdd_datas = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/BDD INITIALE.xlsx'
    try:
        classeur_excel_month_exist = openpyxl.load_workbook(nom_excel_month)
        feuille_excel_month_exist = classeur_excel_month_exist.active
    except FileNotFoundError:
        classeur_exist = None
        
    try:
        classeur_bdd_datas_exist = openpyxl.load_workbook(nom_exel_bdd_datas)
        feuille_bdd_datas_exist = classeur_bdd_datas_exist.active
    except FileNotFoundError:
        classeur_exist = None
        
    #Je charge tous les prompts pour les NFT déjà réalisés
    for nom_nft in os.listdir(dossier_NFT):
        nom_nft = nom_nft[:-4]  # Retirer l'extension ".png"
        winners_with_nft_list.append(nom_nft)
            
    #Je charge l'Excel
    classeur_month = openpyxl.load_workbook(nom_excel_month)
    classeur_bdd_datas = openpyxl.load_workbook(nom_exel_bdd_datas)

    #Ajouts du nom des feuilles Excel présentes dans BDD INITIALE
    sport_sheet = classeur_bdd_datas["SPORT"]
    competition_sheet = classeur_bdd_datas["COMPETITION"]
    event_sheet = classeur_bdd_datas["EVENT"]
    ignore_event_sheet = classeur_bdd_datas["IGNORE EVENT"]
    country_sheet = classeur_bdd_datas["COUNTRY"]
    city_sheet = classeur_bdd_datas["CITY"]
    date_sheet = classeur_bdd_datas["DATE"]
    abreviation_sheet = classeur_bdd_datas["ABREVIATION"]
    comp_of_sport_sheet = classeur_bdd_datas["COMP OF SPORT"]
    month_sheet = classeur_bdd_datas["MONTH"]
    twitter_sheet = classeur_bdd_datas["TWITTER"]
        
    #Ajouts du nom de la feuille Excel présentes dans l'Excel du mois
    ignore_month_elements_sheet = classeur_month["IGNORE_MONTH_ELEMENTS"]


    #CONCERNANT LES ELEMENTS PRESENTS DANS "BDD INITIALE"------------------------------------------------------------------------------
    #----------------------SPORTS EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Sports = [cell.value for cell in sport_sheet['A'] if cell.value is not None]
    EN_Sports = [cell.value for cell in sport_sheet['B'] if cell.value is not None]

    #----------------------COMPETITIONS EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Competition = [cell.value for cell in competition_sheet['A'] if cell.value is not None]
    EN_Competition = [cell.value for cell in competition_sheet['B'] if cell.value is not None]

    #----------------------EVENT EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Event = [cell.value for cell in event_sheet['A'] if cell.value is not None]
    EN_Event = [cell.value for cell in event_sheet['B'] if cell.value is not None]

    #----------------------PAYS EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Country = [cell.value for cell in country_sheet['A'] if cell.value is not None]
    EN_Country = [cell.value for cell in country_sheet['B'] if cell.value is not None]

    #----------------------VILLES EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_City = [cell.value for cell in city_sheet['A'] if cell.value is not None]
    EN_City = [cell.value for cell in city_sheet['B'] if cell.value is not None]

    #----------------------DATES EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Date = [cell.value for cell in date_sheet['A'] if cell.value is not None]
    EN_Date = [cell.value for cell in date_sheet['B'] if cell.value is not None]

    #----------------------ABREVIATIONS EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Abreviation = [cell.value for cell in abreviation_sheet['A'] if cell.value is not None]
    EN_Abreviation = [cell.value for cell in abreviation_sheet['B'] if cell.value is not None]
    ISO3_Abreviation = [cell.value for cell in abreviation_sheet['C'] if cell.value is not None]

    #----------------------EVENTS IGNORES EN ENTREE----------------------#
    # Je créé deux listes plus spécifiques contenant les éléments de la colonne A (en FR) et B (en ANG)
    FR_Ignore_event = [cell.value for cell in ignore_event_sheet['A'] if cell.value is not None]
    
    #----------------------COMPETITION OF SPORT EN ENTREE----------------------#
    #De manière bête et méchante je vais ajouter dans cette liste {competition} of {sport}
    competition_of_sport_list = [cell.value for cell in comp_of_sport_sheet['A'] if cell.value is not None]

    #J'ai ici la traduction de la liste précédente pour que la traduction soit la plus exacte possible
    competition_of_sport_traduction_list = [cell.value for cell in comp_of_sport_sheet['B'] if cell.value is not None]
    


    #----------------------MOIS EN ENTREE----------------------#
    #Traduction du mois afin de faciliter la création des tags produit pour WP
    month_fr_list = [cell.value for cell in month_sheet['A'] if cell.value is not None]
    
    #Traduction du mois afin de faciliter la création des tags produit pour WP
    month_en_list = [cell.value for cell in month_sheet['B'] if cell.value is not None]
    
    
    #----------------------DONNEES TWITTER----------------------#
    #Concernant les gagnants -----------------------------------#
    #Je récupère les gagnants
    winner_twitter_list = [cell.value for cell in twitter_sheet['A'] if cell.value is not None]
    
    #Je récupère les comptes Twitter
    winner_arobase_twitter_list = [cell.value for cell in twitter_sheet['B'] if cell.value is not None]
    
    
    #Concernant les compétitions -----------------------------------#
    #J'ai ici l'arobase de la compétition of sport
    competition_of_sport_arobase_list = [cell.value for cell in comp_of_sport_sheet['D'] if cell.value is not None]
    
    #J'ai ici l'arobase de la compétition of sport
    competition_of_sport_hashtag_list = [cell.value for cell in comp_of_sport_sheet['E'] if cell.value is not None]
    
    #Concernant les events -----------------------------------------#
    #J'ai ici l'arobase des events
    events_arobase_list = [cell.value for cell in event_sheet['C'] if cell.value is not None]
    
    #J'ai ici l'arobase des events
    events_hashtag_list = [cell.value for cell in event_sheet['D'] if cell.value is not None]



    #CONCERNANT LES ELEMENTS PRESENTS DANS l'Excel du mois. IME car "IGNORE_MONTH_ELEMENTS"------------------------------------------------------------------------------
    IME_country_list = [cell.value for cell in ignore_month_elements_sheet['A'] if cell.value is not None]
    IME_city_list = [cell.value for cell in ignore_month_elements_sheet['B'] if cell.value is not None]
    IME_competition_list = [cell.value for cell in ignore_month_elements_sheet['C'] if cell.value is not None]
    IME_comp_of_sport_list = [cell.value for cell in ignore_month_elements_sheet['D'] if cell.value is not None]
    IME_events_list = [cell.value for cell in ignore_month_elements_sheet['E'] if cell.value is not None]
    IME_just_men_woman_list = [cell.value for cell in ignore_month_elements_sheet['F'] if cell.value is not None]
    IME_no_event_probably_empty_list = [cell.value for cell in ignore_month_elements_sheet['G'] if cell.value is not None]
    IME_no_date_event_list = [cell.value for cell in ignore_month_elements_sheet['H'] if cell.value is not None]
    IME_multiple_winnings_same_day_list = [cell.value for cell in ignore_month_elements_sheet['I'] if cell.value is not None]
    IME_no_winner_list = [cell.value for cell in ignore_month_elements_sheet['J'] if cell.value is not None]
    IME_mixed_event_list = [cell.value for cell in ignore_month_elements_sheet['M'] if cell.value is not None]


    IME_ignore_cards_list = [cell.value for cell in ignore_month_elements_sheet['K'] if cell.value is not None]


    if scrapping_month in month_fr_list:
        k = month_fr_list.index(scrapping_month)
        month_eng = month_en_list[k]
    else:
        print('SOUCIS DE TRADUCTION MOIS')

    #-------------------------------------------------FIN ELEMENTS DE VERIFICATION EN ENTREE-------------------------------------------------#
    #----------------------------------------------------------------------------------------------------------------------------------------#

    #Appel de mon url contenant les données
    result = requests.get(url)

    if result.status_code == 200:
        doc = BeautifulSoup(result.text, "html.parser")
        all_months = doc.select('li a.toggle-btn')
        for month in all_months:
            #print(month)
            if month.text == scrapping_month:
                events = month.find_next('div')
                all_events = events.select('tr')
                for event in all_events:
                    
                    if EVENT_COUNTER<1000 :
                        
                        event_informations = event.select('td')
                        if event_informations :
                            COMPETITION_COUNTER +=1
                            
                            
    #---------------------------------------------------------------------------------------------------------------------------------------------------Vérification des infos présentes dans le tableau principal
    #-----------------------------------------------------------------------------------On vérifie la bonne traduction des Pays, Villes, Sports, Compétitions

                        
                            #Recherche dans la première partie du tableau initial (dates)
                            competition_date = event_informations[0].text if len(event_informations) > 0 else None
                            
                            #Je vais directement exclure de mon scrapping les compétitions non terminées au jour d'aujourd'hui
                            check_competition_date_end = competition_date
                            if isinstance (check_competition_date_end, list):
                                pass
                            else :
                                if "-" in check_competition_date_end:
                                    check_competition_date_end = check_competition_date_end.split("-")
                                    competition_end_date = check_competition_date_end[1]
                                else:
                                    competition_end_date = check_competition_date_end
                            
                            if scrapping_month in competition_end_date :
                                competition_end_date = competition_end_date.replace(actual_year, "")
                                competition_end_date_day = int(''.join(filter(str.isdigit,competition_end_date)))
                                if competition_end_date_day < actual_day :
                                    
                                    #Recherche dans la deuxième partie du tableau initial (Pays/ville)
                                    competition_country = event_informations[1].text if len(event_informations) > 0 else None
                                    competition_city_first_chance = event_informations[1].text if len(event_informations) > 0 else None
                                    
                                    #Recherche dans la troisième partie du tableau initial (sport/épreuve/compétition/ville)
                                    sport = event_informations[2].text if len(event_informations) > 0 else None
                                    sport_competition = event_informations[2].text if len(event_informations) > 0 else None
                                    competition_city_second_chance = event_informations[2].text if len(event_informations) > 0 else None


            #-----------------------On va chercher un pays précis dans la 2ème colonne du tableau principal
                                    if competition_country :
                                        Country_match = [Country for Country in FR_Country if Country.lower() in competition_country.lower()]
                                        if Country_match:
                                            Good_country = max(Country_match, key=len)
                                            if Good_country in FR_Country:
                                                Good_country_fr = Good_country
                                                index_pays = FR_Country.index(Good_country)
                                                competition_country = EN_Country[index_pays]
                                        else :
                                            no_country_list.append(competition_country)
                                            competition_country = None
                                    else :
                                        print("Aucune chance d'arriver ici. Pas de valeur dans la 2ème colonne du tableau principal")

            #-----------------------On va chercher une ville précise dans la 2ème et 3ème colonne du tableau principal
                                    if competition_city_first_chance and competition_city_second_chance :
                                        City_first_chance_match = [City for City in FR_City if City.lower() in competition_city_first_chance.lower()]
                                        City_second_chance_match = [City for City in FR_City if City.lower() in competition_city_second_chance.lower()]

                                        if City_first_chance_match :
                                            Good_city = max(City_first_chance_match, key=len)
                                            if Good_city in FR_City:
                                                index_city = FR_City.index(Good_city)
                                                city = EN_City[index_city]
                                        elif City_second_chance_match :
                                            Good_city = max(City_second_chance_match, key=len)
                                            if Good_city in FR_City:
                                                index_city = FR_City.index(Good_city)
                                                city = EN_City[index_city]
                                        else:
                                            no_city_list.append(f"{competition_city_first_chance} ou {competition_city_second_chance}")
                                            city = "" #Possible qu'on ait juste pas de ville. Je mets une valeur vide mais mets quand même une alerte pour la fin
                                            
                                        if city == "-":
                                            city = ""
                                    else:
                                        print("Aucune chance d'arriver ici. Pas de valeur dans la 2ème ou 3ème colonne du tableau principal")


            #-----------------------On va chercher un sport précise dans la 3ème colonne du tableau principal
                                    if sport :
                                        sport_matches = [Sport for Sport in FR_Sports if Sport.lower() in sport.lower()]
                                        if sport_matches:
                                            Good_sport = max(sport_matches, key=len)
                                            if Good_sport in FR_Sports:
                                                index_sport = FR_Sports.index(Good_sport)
                                                Good_sport_eng = EN_Sports[index_sport]
                                                sport = Good_sport_eng.strip()
                                        else:
                                            no_sport_list.append(sport)
                                            sport = None
                                    else:
                                        print("Aucune chance d'arriver ici. Pas de valeur dans la 3ème colonne du tableau principal")


            #-----------------------On va chercher une compétition précise dans la 3ème colonne du tableau principal
                                    if sport_competition :
                                        sport_competition = sport_competition.lower()
                                        competition_matches = [Competition for Competition in FR_Competition if Competition.lower() in sport_competition.lower()]
                                        if competition_matches:
                                            Good_competition = max(competition_matches, key=len)
                                            if Good_competition in FR_Competition:
                                                index_competition = FR_Competition.index(Good_competition)
                                                Good_competition_eng = EN_Competition[index_competition]
                                                sport_competition = Good_competition_eng
                                        else:
                                            no_sport_competition_list.append(sport_competition)
                                            sport_competition = ""
                                    else:
                                        print("Aucune chance d'arriver ici. Pas de valeur dans la 3ème colonne du tableau principal")       



            #---------------------------------------------------------------------------------------------------------------------------------------------------Vérification des infos présentes dans les pages d'events
            #-----------------------------------------------------------------------------------

                                    sport_event_link = event_informations[2].select_one('a')['href'] if len(event_informations) > 2 and event_informations[2].select_one('a') else None
                                    sport_event_link_text = "https://www.les-sports.info/" + sport_event_link if sport_event_link else None
                                    
                                    if "hommes-epm" in sport_event_link_text :
                                        url_event_hommes = sport_event_link_text
                                        url_event_femmes = sport_event_link_text.replace("hommes-epm", "femmes-epf")
                                        url_event_mixte = sport_event_link_text.replace("hommes-epm", "mixte-epi")
                                        mixte_event = True
                                    else:
                                        url_event_hommes = sport_event_link_text
                                        url_event_femmes = ""
                                        url_event_mixte = ""
                                        mixte_event = False
                                    urls_event_list = [url_event_hommes, url_event_femmes, url_event_mixte]
                                    
                                    for url_event in urls_event_list :
                                        if "mixte-epi" in url_event :
                                            mixte_event = True
                                        else :
                                            mixte_event = False
                                        if url_event != None and url_event != "" :
                                            event_detail = requests.get(url_event)

                                            if result.status_code == 200:
                                                all_competition = BeautifulSoup(event_detail.text, "html.parser")


                                                #Je cherche les gagnants dans la page d'event pour la BOUCLE 1
                                                team_competition = None
                                                team_competition = all_competition.select(".h3-vainqueur") #Si j'ai une valeur dans cette classe, c'est que je suis dans une compétition par équipe (1 compétition, 1 gagnant)
                                                
                                                #Je cherche les titres des tableaux de résultats Pour la BOUCLE 2
                                                events_ok_list.clear()
                                                all_events_in_page = None #Remise à zéro
                                                all_events_in_page = all_competition.select(".nomargin a") #Si j'ai une valeur dans cette classe ou la suivante, je suis dans une compétition qui a plusieurs events donc plusieurs gagnants
                                                if all_events_in_page :
                                                    pass
                                                else:
                                                    all_events_in_page = all_competition.select(".tab_content h2.centre")
                                                
            #----------------------------------------------------------------------------------------BOUCLE1---------------------------------------------------------------------------------
            #-------------------------------------------------------------------------------Vainqueur = Equipe nationale---------------------------------------------------------------------                                 
                                                if team_competition :
                                                    for winning_team in team_competition:
                                                        if winning_team :
                                                            EVENT_COUNTER +=1
                                                            sport_event = "" #Pas d'event particulier, l'équipe gagne la compétition en elle-même
                                                            date_event = "" #Avoir la date de la finale ne m'intéresse pas vu que l'équipe gagne une compétition de plusieurs jours
                                                            winner_country = "" #C'est déjà un pays qui gagne
                                                            city = "" #Les compétitions par équipe se passent souvent dans plusieurs villes
                                                            
                                                            winning_team_info = winning_team.find(class_='nodecort')
                                                            winning_team_name = winning_team_info['title'] if winning_team_info is not None else None
                                                            
                                                            if winning_team_name in FR_Country:
                                                                index_country = FR_Country.index(winning_team_name)
                                                                winner = EN_Country[index_country]
                                                                
                                                                competition_of_sport = f"{sport_competition} of {sport}"
                                                                if competition_of_sport in competition_of_sport_list:
                                                                    j = competition_of_sport_list.index(competition_of_sport)
                                                                    competition_of_sport_traduction_value = competition_of_sport_traduction_list[j]
                                                                    
                                                                    if competition_country :
                                                                        prompt_initial = f'{winner} wins the {competition_of_sport_traduction_value} in {competition_country}'
                                                                    else :
                                                                        prompt_initial = f'{winner} wins the {competition_of_sport_traduction_value}' #Le tournoi des VI Nations par ex se passe dans plusieurs pays
                                                                    
                                                                    winner_len = str(len(winner))
                                                                    name_NFT = card_name_without_accent(f"{winner_len}-{sport}-{sport_competition}-{actual_year}")

                                                                    #J'intègre d'abord TOUS les évènements du mois dans ma feuille excel ALL
                                                                    all_winners_one_sheet = dictionary.add_to_ALL_sheet(competition_date,competition_country,city,sport,sport_competition,sport_event,date_event,winner,winner_country,url_event,prompt_initial,actual_year,name_NFT)
                                                                    
                                                                    #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la page contenant TOUTES les compétitions du mois
                                                                    all_month_winners_list.append(all_winners_one_sheet)
                                                                    
                                                                    if name_NFT not in winners_with_nft_list : #La carte n'est pas encore créée. j'envoi ces données dans la liste du jour
                                                                        #Ci-dessous les éléments spécifiques à intégrer à la feuille contenant les évènements n'ayant pas encore de carte
                                                                        EVENT_SPECIFIC_COUNTER +=1
                                                                        winner_tweet = ""
                                                                        new_winners_one_sheet = dictionary.add_to_today_sheet(EVENT_SPECIFIC_COUNTER,competition_date,competition_country,city,sport,sport_competition,sport_event,date_event,winner,winner_country,url_event, prompt_initial,actual_year,name_NFT,winner_tweet)
                                                                        rename_prompt_for_midjourney(prompt_initial) #Le prompt devient le nom sous lequel j'enregistre rapidement chaque image Midjourney. J'ajoute la clé {Prompt_Midjourney} à new_winners_one_sheet
                                                                        
                                                                        #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la page du jour qui contient tous les events sans cartes
                                                                        winners_without_nft_list.append(new_winners_one_sheet)
                                                                        
                                                                        #Ci-dessous les éléments spécifiques à intégrer à la feuille qui permet l'import des produits dans WP
                                                                        prompt_for_import_product = name_NFT + ".png"
                                                                        short_winner = create_short_winner(winner) #Sert pour réduire la taille du titre de la carte sur le site
                                                                        
                                                                        #Je dois récupérer le mois de l'event et le mettre dans le tag (et non pas mettre le mois du scrapping). Certaines compétitions commencent en fev et finissent en mars par ex
                                                                        #Pour les events par équipe, je prends la date de la compétition et je chope le mois dans la fin de variable
                                                                        competition_date = competition_date.split("-")
                                                                        competition_end_date = competition_date[1]
                                                                        for month in month_fr_list :
                                                                            if month in competition_end_date :
                                                                                k = month_fr_list.index(month)
                                                                                month_event = month_en_list[k]
                                                                            
                                                                            
                                                                        data_for_wordpress = dictionary.import_wordpress (EVENT_COUNTER,short_winner,winner,sport,sport_competition,sport_event,prompt_for_import_product,actual_year,prompt_initial,month_eng,name_NFT,month_event)
                                                                        
                                                                        #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la liste qui servira à compléter l'Excel à la date du scrapping
                                                                        data_for_wordpress_list.append(data_for_wordpress)
                                                                    
                                                                else: #Pas de traduction pour "competition of sport" donc les infos ne sont pas intégrées à l'excel
                                                                    no_competition_of_sport_list.append(f'{sport_competition} of {sport} - {url_event}') #J'ajoute un message final pour rajouter comp of sport dans ma liste Excel
                                                                    
                                                            else: #Pas de traduction pour le pays gagnant donc les infos ne sont pas intégrées à l'excel
                                                                no_country_list.append(f"{winning_team_name} - {url_event}")

                                                        else:
                                                            no_winner_list.append(f"BALISE_no_winner 4 : {url_event}")
                                                
                                                
                                                
            #----------------------------------------------------------------------------------------BOUCLE2---------------------------------------------------------------------------------
            #-------------------------------------------------------------------------------Vainqueur = Individu(s)--------------------------------------------------------------------------                                    
                                                elif all_events_in_page :
                                    #---------------1er tris : je retire tous les events présents dans ma liste "Ignore_Competitions"
                                                    for event_in_page in all_events_in_page :
                                                        if event_in_page :
                                                            event_in_page_text = event_in_page.text
                                                            if [Ignore_Competition for Ignore_Competition in FR_Ignore_event if Ignore_Competition.lower() in event_in_page_text.lower()]:
                                                                pass #L'event est ignoré, je ne fais rien de plus avec lui car dans la suite je ne vais itérer que dans la liste [events_ok_list]
                                                            else:
                                                                events_ok_list.append(event_in_page_text) #Les events non ignorés sont ajoutés à cette liste events_ok_list
                                                                
                                    #---------------2ème tris : Parmis les éléments qui ont passé le premier tamis et qui sont dans la liste d'events de la compet, je vais maintenant exclure ceux situés avant une class p_14 indiquant un event annulé
                                                                p_14_elements = all_competition.find_all(class_="p_14") #Je cherche toutes ces classes p_14
                                                                for p_14_element in p_14_elements:
                                                                    if p_14_element and re.search(r"Cette manche a été annulée", p_14_element.text, re.IGNORECASE):
                                                                        event_canceled = p_14_element.find_previous_sibling(class_="nomargin") #l'event annulé est situé dans la class nomargin située juste avant
                                                                        event_canceled = event_canceled.text
                                                                        if event_canceled == event_in_page_text : #si l'event avant p_14 est l'event dans lequel on est, je l'enlève de la liste [events_ok_list]
                                                                            events_ok_list.remove(event_canceled)
                                                                        else:
                                                                            pass 
                                                                    else:
                                                                        pass #Pas de manche annulée, possible que ce p_14 serve à autre chose
                                                                    
                                    #---------------3ème tris : je retire les events présents dans un toggle (résultats détaillés donc donnée en double)
                                                                if event_in_page.find_parent("ul", class_="toggle"): #si un event a une class toggle en parent c'est qu'il doit être remove de [events_ok_list]
                                                                    events_ok_list.remove(event_in_page_text)
                                                                else:
                                                                    pass #L'event n'est pas dans un toggle "Résultat détaillé". Je le laisse dans ma liste [events_ok_list]
                                                        else:
                                                            no_event_list.append(f"BALISE_no_event 4 : {url_event}")

                                                    #J'ai identifié les events à prendre en compte
                                                    for event_in_page_index, event_in_page in enumerate(all_events_in_page,start=1): #Je refais un for sur tous les events de la page
                                                        if event_in_page.text in events_ok_list : #Je vérifie par contre tout de suite s'il fait partie de la liste [events_ok_list] pour ne continuer qu'avec les events passés par les tamis précédents
                                                            sportsmen_table = event_in_page.find_all_next('table', class_='table-style-2', limit=1) #Je chope tous les tableaux contenant les athlètes
                                                            specific_event_title = event_in_page.text
                                                            
                                                            #Je vérifie si dans specific_event_title je retrouve un nom d'event présent dans ma BDD INITIALE
                                                            event_matches = [Event for Event in FR_Event if Event.lower() in specific_event_title.lower()]
                                                            if event_matches: #Si l'épreuve fait partie de la liste des events dans BDD INITIALE
                                                                Good_event = max(event_matches, key=len) #Je prends la plus longue correspondance s'il y en a plusieurs
                                                                index_event = FR_Event.index(Good_event)
                                                                sport_event = EN_Event[index_event] #sport_event a maintenant sa version traduite en anglais
                                              
                                                                
                                                                if len(Good_event) == 2: #J'ai créé 1 event "ho" et 1 event "fe" (donc si ==2 c'est qu'on a que homme ou femme comme info pour l'event ou bien que l'event n'est pas traduit)
                                                                    just_men_woman_list.append(f"BALISE_no_event 3 : {specific_event_title} - {url_event}") #Je permets de voir rapidement dans le print final si je n'ai que Homme/Femme comme info pour l'event
                                                            
                                                            else: #L'épreuve ne fait pas partie de ma BDD INITIALE.
                                                                events_ok_list.remove(specific_event_title) #Je retire cet event de la liste des events ok, il devra être traduit si on veut son ajout dans l'Excel
                                                                no_event_list.append(f"BALISE_no_event 2 : {specific_event_title} - {url_event}")
                                                                sport_event = "" #Je mets une valeur vide afin de ne pas avoir de confusion dans l'Excel
                                                                
                                                            # if sport_event in EN_Event :
                                                            #     index_event = EN_Event.index(sport_event)
                                                            #     event_arobase = events_arobase_list[index_event]
                                                            #     event_hashtag = events_arobase_list[index_event]
                                                            # else:
                                                            #     event_arobase = ""
                                                            #     event_hashtag = ""
                                    
                                                            #Je prends le titre d'event et je cherche une date présente dans ma BDD INITIALE
                                                            date_matches = [Date for Date in FR_Date if Date.lower() in specific_event_title.lower()]
                                                            if date_matches:
                                                                Good_date = max(date_matches, key=len)
                                                                index_date = FR_Date.index(Good_date)
                                                                date_event = EN_Date[index_date] #Date a maintenant sa version traduite en anglais
                                                            else:
                                                                no_date_event_list.append(f"BALISE_no_date 3 : {specific_event_title} - {competition_date} {url_event}")
                                                                
                                                            if sportsmen_table :
                                                                solo_winner = None #Je remets à zéro au cas où
                                                                team_winner = None #Je remets à zéro au cas où
                                                                
                                                                first_row_infos = sportsmen_table[0]
                                                                solo_winner = first_row_infos.find('a', class_='nodecort') #Si j'ai un gagnant seul, class = nodecort
                                                                team_winner = first_row_infos.find(class_='tdcol-70') #Si j'ai plusieurs gagnants, class = tdcol-70
                                                                
                                                                if solo_winner :  
                                                                    winner_style = "solo"
                                                                    winner = (solo_winner.text).strip() #Je retire tous les espaces avant/après
                                                                elif team_winner :
                                                                    winner_style = "team"
                                                                    winner = (team_winner.text).strip() #Je retire tous les espaces avant/après
                                                                else:
                                                                    date_number = re.search(r'\d+', date_event)
                                                                    date_number_int = int(date_number.group())
                                                                    if date_number_int > verif_date_event :
                                                                        #La date de l'event est supérieur à la date de vérif. Je ne print rien. Il y a un tableau sans gagnant mais normal de ne pas avoir de gagnant donc...
                                                                        winner = None
                                                                    elif date_number_int < verif_date_event :
                                                                        winner = None
                                                                        no_winner_list.append(f"BALISE_no_winner 3 : {url_event}") #La date de l'event est antérieur à la date du scrapping. Je vérifie l'url pour voir pourquoi on a pas de gagnant
                                                                    else:
                                                                        no_date_event_list.append(f"BALISE_no_date 2 : {url_event}") #J'ai peut-être un soucis avec le if et elif. Pas de winner et soucis de date ? A checker

                                                                #résultat sous la forme "nom (pays)" si "solo" ou "pays (nom1, nom2)" si "team"
                                                                if winner is not None :
                                                                    winners_info = winner.split('(')
                                                                    if len (winners_info) > 1:
                                                                        info_1 = winners_info[0].strip()
                                                                        info_2 = winners_info[1].split(')')[0].strip()
                                                                        if winner_style == "solo":
                                                                            winner = info_1
                                                                            winner_country = info_2
                                                                        else:
                                                                            winner = info_2
                                                                            winner_country = info_1
                                                                        
                                                                        #On va traduire le pays pour qu'il corresponde à la norme ISO 3 et ainsi obtenir un prompt propre et uniformisé
                                                                        #url source : https://www.trucsweb.com/tutoriels/internet/iso_3166/
                                                                        if winner_country in ISO3_Abreviation: #Si l'abréviation répond déjà à ISO 3
                                                                            winner_country_info = True
                                                                        elif winner_country in FR_Abreviation: #Si l'abréviation est présente en Français, on choisit sa traduction en ISO 3
                                                                            index_FR_Abr = FR_Abreviation.index(winner_country)
                                                                            winner_country = EN_Abreviation[index_FR_Abr]
                                                                            winner_country_info = True
                                                                        else:
                                                                            no_abr_list.append(f"BALISE_no_abr 1 : {winner_country}")
                                                                            winner_country = "?"
                                                                            winner_country_info = False
                                                                    else:
                                                                        winner = winners_info[0]
                                                                        winner_country = "?"
                                                                        winner_country_info = False
                                                                        #Trouver un moyen à terme de récupérer la nationalité de l'athlète (en attendant prompt sans nationalité)
                                                                    
                                                                    #J'ai le gagnant et la date de sa victoire. Je vérifie si je n'ai pas des athlètes ayant gagné plusieurs fois le même jour (soucis sur le site potentiel)
                                                                    winner_event_date_concordance(winner,date_event,url_event)
                                                                    
                                                                    EVENT_COUNTER +=1 #J'ai passé toutes les embûches, à partir d'ici je vais organiser les données pour les inclure dans les feuilles Excel
                                                                    sport_competition = sport_competition.strip()
                                                                    sport = sport.strip()
                                                                    competition_of_sport = f"{sport_competition} of {sport}"
                                                                    
                                                                    
                                                                    #Je vais dans le if/else suivant ressortir le prompt_initial
                                                                    if competition_of_sport in competition_of_sport_list:
                                                                        competition_of_sport_index = competition_of_sport_list.index(competition_of_sport)
                                                                        competition_of_sport_traduction_value = competition_of_sport_traduction_list[competition_of_sport_index]
                                                                        
                                                                        
                                                                        
                                                                        if winner_country_info :
                                                                            if city == "": #J'ai l'info de la nationalité du gagnant mais pas de ville d'épreuve
                                                                                prompt_initial = f'{winner} ({winner_country}) wins the {sport_event} {competition_of_sport_traduction_value} in {competition_country} on {date_event}'
                                                                            else: #J'ai l'info de la nationalité du gagnant mais et la ville d'épreuve
                                                                                prompt_initial = f'{winner} ({winner_country}) wins the {sport_event} {competition_of_sport_traduction_value} in {city}, {competition_country} on {date_event}'
                                                                        else :
                                                                            if city == "": #Je n'ai pas l'info de la nationalité du gagnant ni la ville d'épreuve
                                                                                prompt_initial = f'{winner} wins the {sport_event} {competition_of_sport_traduction_value} in {competition_country} on {date_event}'
                                                                            else: #Je n'ai pas l'info de la nationalité du gagnant mais j'ai la ville d'épreuve
                                                                                prompt_initial = f'{winner} wins the {sport_event} {competition_of_sport_traduction_value} in {city}, {competition_country} on {date_event}'
                                                                    else:
                                                                        competition_of_sport_index = 0
                                                                        prompt_initial = 'Pas de prompt' #Je dois avoir une formulation de prompt nickel
                                                                        no_competition_of_sport_list.append(f'{sport_competition} of {sport} - {url_event}')
                                                                        
                                                                    prompt_initial = prompt_initial.replace("  ", " ") #J'ai des cas où je n'ai pas d'event (tennis par ex ou je n'arrive pas à distinguer si ATP ou WTA). Je transforme le double espace en simple
                                                                    
                                                                    if mixte_event == True :
                                                                        mixed_event_list.append(sport_event) #Je vérifie si la mixité de l'épreuve est bien dans l'event
                                                                        
                                                                        
                                                                    #J'ai le prompt_initial et toutes les infos, je créé le nom de la carte et j'ajoute les données à l'Excel
                                                                    winner_len = len(winner)
                                                                    name_NFT = f"{winner_len}-{sport}-{sport_competition}-{sport_event}-{date_event}-{actual_year}"
                                                                    name_NFT = card_name_without_accent(name_NFT)
                                                                    
                                                                    #J'intègre d'abord TOUS les évènements du mois dans ma feuille excel ALL
                                                                    all_winners_one_sheet = dictionary.add_to_ALL_sheet(competition_date,competition_country,city,sport,sport_competition,sport_event,date_event,winner,winner_country,url_event, prompt_initial,actual_year,name_NFT)
                                                                    #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la page contenant TOUTES les compétitions du mois
                                                                    all_month_winners_list.append(all_winners_one_sheet) #j'ajoute le dictionnaire à ma liste contenant tous les gagnants et leurs infos annexes
                                                                    
                                                                    #Je vais éviter d'avoir des résultats préliminaires pour des compétitions non terminées encore
                                                                    


                                                                    if name_NFT not in winners_with_nft_list : #La carte n'est pas encore créée. j'envoi ces données dans la liste du jour
                                                                        if name_NFT not in IME_ignore_cards_list :
                                                                            #Ci-dessous les éléments spécifiques à intégrer à la feuille contenant les évènements n'ayant pas encore de carte
                                                                            EVENT_SPECIFIC_COUNTER +=1
                                                                            
                                                                            #Je m'occupe des données relatives à Twitter
                                                                            twitter_results = twitter_datas(competition_of_sport_index, winner)
                                                                            arobase_competition_value = twitter_results['arobase']
                                                                            hashtag_competition_value = twitter_results['hashtag']
                                                                            winner_twitter_account = twitter_results['winner_account']
                                                                            
                                                                            winner_tweet=f"Hey {winner_twitter_account}, don't forget to retweet us to get your free winning card !! {arobase_competition_value} {hashtag_competition_value}. 1$ for everyone else !"
                                                                            winner_tweet = winner_tweet.replace("None","")
                                                                            
                                                                            new_winners_one_sheet = dictionary.add_to_today_sheet(EVENT_SPECIFIC_COUNTER,competition_date,competition_country,city,sport,sport_competition,sport_event,date_event,winner,winner_country,url_event, prompt_initial, actual_year,name_NFT,winner_tweet)
                                                                            rename_prompt_for_midjourney(prompt_initial)
                                                                            
                                                                            #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la page du jour qui contient tous les events sans cartes
                                                                            winners_without_nft_list.append(new_winners_one_sheet)
                                                                            
                                                                            #Ci-dessous les éléments spécifiques à intégrer à la feuille qui permet l'import des produits dans WP
                                                                            prompt_for_import_product = name_NFT + ".png"
                                                                            short_winner = create_short_winner(winner) #Sert pour réduire la taille du titre de la carte sur le site
                                                                            
                                                                            #Je dois récupérer le mois de l'event et le mettre dans le tag (et non pas mettre le mois du scrapping). Certaines compétitions commencent en fev et l'event a lieu en mars par ex
                                                                            #Pour les events en indiv, je prends la date de l'event et je chope le mois dedans
                                                                            for month in month_en_list :
                                                                                if month in date_event :
                                                                                    month_event = month
                                                                            
                                                                            data_for_wordpress = dictionary.import_wordpress (EVENT_COUNTER,short_winner,winner,sport,sport_competition,sport_event,prompt_for_import_product,actual_year,prompt_initial,month_eng,name_NFT,month_event)
                                                                            
                                                                            #J'ai toutes les valeurs pour l'Excel, j'envoi les données du dictionnaire vers la liste qui servira à compléter l'Excel à la date du scrapping
                                                                            data_for_wordpress_list.append(data_for_wordpress)

                                                                        else : 
                                                                            cards_ignored_list.append(f"{winner} - {name_NFT}")
                                                                else:
                                                                    no_winner_list.append(f"BALISE_no_winner 2 : {url_event}")
                                                                
                                                            else:
                                                                date_number = re.search(r'\d+', date_event)
                                                                date_number_int = int(date_number.group())
                                                                if date_number_int > actual_day :
                                                                        #L'event n'a pas encore eu lieu. Pas de soucis et je ne vais pas plus loin dans la recherche, je pass pour cet event
                                                                        winner = None
                                                                elif date_number_int < actual_day :
                                                                    winner = None
                                                                    #La date de l'event est antérieur à la date du scrapping. Potentiel soucis. J'ajoute l'info dans le print final et m'arrête là dans les recherches
                                                                    no_winner_list.append(f"BALISE_no_winner 1 : {competition_date} - {url_event}")
                                                                else :
                                                                    no_date_event_list.append(f"BALISE_no_date 1 : {url_event}") #J'ai peut-être un soucis avec le if et elif. Pas de winner et soucis de date ? A checker
                                                                    
                                                        else :
                                                            pass #L'event ne fait pas partie de [events_ok_list], j'arrive ici et je pass

                                                    pause = random.randrange(1, 3)
                                                    time.sleep(pause)   

                                                else:
                                                    no_event_probably_empty_list.append(f"BALISE_no_event 1 : {competition_date} - {url_event}") #J'ai une page d'event mais pas de gagnant ni présent dans h3, no margin ou center
                                            else:
                                                print(f"Pas de retour de l'url. Status code = {result.status_code}")

                                    
        #----------------------------------------------------------------------------------------PLACE A LA CREATION DE L'EXCEL---------------------------------------------------------------------------------
                        #-----------j'ajoute mon dictionnaire dans ma liste d'events
                                    #competition_list.append(competition_dict)
                                    
                                else:
                                    pass #Je passe ici si par exemple je n'ai pas d'url pour les femmes et/ou en mixte
                                
            else:
                pass #On fait quoi si on est pas au actual month ? Bah rien on sort de la boucle et on arrête


    #J'imprime en fin de scrapping toutes les erreures ensembles par catégorie afin de faciliter la lecture
        print()
        print(f"Informations pour le {actual_day} {scrapping_month}")
        print()
        print()
        print()
        print()

        if no_country_list :
            print("\033[4m" + 'Manque les pays suivants dans COUNTRY. Ajouter les lignes à ignorer en colonne A de IME :' + "\033[0m", end="")
            print()
            print()
            for no_country in no_country_list:
                if no_country in IME_country_list:
                    pass
                else:
                    print(f"{no_country}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_city_list :
            print("\033[4m" + 'Manque les villes suivantes dans CITY. Ajouter les lignes à ignorer en colonne B de IME' + "\033[0m", end="")
            print()
            print()
            for no_city in no_city_list:
                no_city = no_city.strip()
                if no_city in IME_city_list:
                    pass
                else:
                    print(f"{no_city}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_sport_list :
            print("\033[4m" + 'Manque les sports suivants dans SPORT :' + "\033[0m", end="")
            print()
            for no_sport in no_sport_list:
                print(f" - {no_sport}")
            print()
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_sport_competition_list :
            print("\033[4m" + 'Manque les compétitions suivantes dans COMPETITION. Ajouter les lignes à ignorer en colonne C de IME' + "\033[0m", end="")
            print()
            print()
            for no_sport_competition in no_sport_competition_list:
                if no_sport_competition in IME_competition_list:
                    pass
                else:
                    print(f"{no_sport_competition}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_competition_of_sport_list :
            print("\033[4m" +'Manque les "compétitions of sport" suivantes dans COMP OF SPORT. Ajouter les lignes à ignorer en colonne D de IME' + "\033[0m", end="")
            print()
            print("\033[4m" +"Ces résultats NE SONT PAS présents dans l'Excel" + "\033[0m", end="")
            print()
            print()
            for no_competition_of_sport in no_competition_of_sport_list:
                if no_competition_of_sport in IME_comp_of_sport_list:
                    pass
                else:
                    print(f"{no_competition_of_sport}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_event_list :
            print("\033[4m" +'Manque les events suivants dans EVENT. Ajouter les lignes à ignorer en colonne E de IME' + "\033[0m", end="")
            print()
            print("\033[4m" +"Ces résultats SONT dans l'Excel donc le prompt n'a pas d'event. A prendre en considération" + "\033[0m", end="")
            print()
            print()
            for no_event in no_event_list:
                if no_event in IME_events_list:
                    pass
                else:
                    print(f"{no_event}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if just_men_woman_list :
            print("\033[4m" +"Probablement que HOMME/FEMME dans l'event. Ajouter infos dans 'EVENT'. Ajouter les lignes à ignorer en colonne F de IME" + "\033[0m", end="")
            print()
            print()
            for just_men_woman in just_men_woman_list:
                if just_men_woman in IME_just_men_woman_list:
                    pass
                else:
                    print(f"{just_men_woman}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_event_probably_empty_list :
            print("\033[4m" +'Pages très probablement vides. Ajouter les lignes à ignorer en colonne G de IME' + "\033[0m", end="")
            print()
            print()
            for no_event_probably_empty in no_event_probably_empty_list:
                if no_event_probably_empty in IME_no_event_probably_empty_list:
                    pass
                else:
                    print(f"{no_event_probably_empty}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_date_event_list :
            print("\033[4m" +"Manque les dates d'event suivantes dans DATE. Ajouter les lignes à ignorer en colonne H de IME" + "\033[0m", end="")
            print()
            print()
            for no_date_event in no_date_event_list:
                if no_date_event in IME_no_date_event_list:
                    pass
                else:
                    print(f"{no_date_event}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_winner_list :
            print("\033[4m" +"Manque un gagnant dans les events suivants. Ajouter les lignes à ignorer en colonne J de IME" + "\033[0m", end="")
            print()
            print()
            for no_winner in no_winner_list:
                if no_winner in IME_no_winner_list:
                    pass
                else:
                    print(f"{no_winner}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if no_abr_list:
            print("\033[4m" + "Manque les abréviations suivantes dans ABREVIATION : " + "\033[0m", end="")
            print()
            print()
            for abr_translation in no_abr_list:
                print(f"{abr_translation}")    
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if multiple_winnings_same_day_list :
            print("\033[4m" + "Ces athlètes ont remporté plusieurs épreuves le même jour. Ajouter les lignes à ignorer en colonne I de IME" + "\033[0m", end="")
            print()
            print()
            for winnings_same_day in multiple_winnings_same_day_list:
                if winnings_same_day in IME_multiple_winnings_same_day_list:
                    pass
                else:
                    print(f"{winnings_same_day}") 
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
        
        if cards_ignored_list :
            print("\033[4m" + "Ces cartes ne seront pas créées, l'event est bien dans ALL. Cartes exclues présentes en colonne J de IME" + "\033[0m", end="")
            print()
            print()
            for cards_ignored in cards_ignored_list :
                print(f"{cards_ignored}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if twitter_account_list :
            print("\033[4m" + "Je ne sais pas si ces gagnants ont un compte Twitter. Ajouter l'info en page TWITTER de la BDD INITIALE" + "\033[0m", end="")
            print()
            print("\033[4m" +"Le Tweet n'est PAS utilisable, pas de gagnant nommé !" + "\033[0m", end="")
            print()
            for twitter_acount in twitter_account_list :
                print(f"{twitter_acount}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if  no_competition_arobase_list:
            print("\033[4m" + "Manque le @ Twitter pour ces COMP OF SPORT" + "\033[0m", end="")
            print()
            print()
            for no_competition_arobase in no_competition_arobase_list :
                print(f"{no_competition_arobase}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if  no_competition_hashtag_list:
            print("\033[4m" + "Manque le # Twitter pour ces COMP OF SPORT" + "\033[0m", end="")
            print()
            print()
            for no_competition_hashtag in no_competition_hashtag_list :
                print(f"{no_competition_hashtag}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if  mixed_event_list:
            print("\033[4m" + "Epreuve mixte. Vérifier si le prompt en tient compte" + "\033[0m", end="")
            print()
            print()
            for mixed_event in mixed_event_list :
                if mixed_event in IME_mixed_event_list:
                    pass
                else:
                    print(f"{mixed_event}")
            print(f'---------------------------------------------------------------------------------------------------------------------------------------------------------------')
            print()
            print()
            print()
            
        if winners_without_nft_list :
            print("\033[4m" + "Voici les prompts pour créer les images sur Midjourney des derniers vainqueurs identifiés " + "\033[0m", end="")
            print()
            print()
            tour=0
            new_prompts = 0
            for winner_without_nft_list in winners_without_nft_list:
                if tour < 14:
                    print()
                    print(f"{winner_without_nft_list['Prompt']} {midjourney_parameters}")
                    #discord_prompt_list.append(f"{winner_without_nft_list['Prompt']} {midjourney_parameters}")
                    tour +=1
                    new_prompts +=1
                else:
                    print()
                    print(f"{winner_without_nft_list['Prompt']} {midjourney_parameters}")
                    #discord_prompt_list.append(f"{winner_without_nft_list['Prompt']} {midjourney_parameters}")
                    tour=0
                    new_prompts +=1
                    print()
                    print("--------------------------------------------------")
            print()
            print(f'--------------------------------------------------')
            print(f"Nous avons {new_prompts} nouveaux prompts")
            print(f'--------------------------------------------------')
            print()
        
        for discord_prompt in discord_prompt_list:
            # Créez le message à envoyer
            message = {
                'content': discord_prompt
            }
            # Envoyez le message au webhook Discord
            response = requests.post(webhook_url, data=json.dumps(message), headers={'Content-Type': 'application/json'})

    #Création de l'Excel   
        actual_day = str(actual_day)
        # Créer un DataFrame pandas à partir de la liste d'événements
        df1 = pd.DataFrame(all_month_winners_list)
        df2 = pd.DataFrame(winners_without_nft_list)
        df3 = pd.DataFrame(data_for_wordpress_list)

        with pd.ExcelWriter(nom_excel_month, engine='openpyxl', mode='a') as writer:
            try:
                writer.book.remove(writer.book["ALL"])
            except KeyError:
                pass
        with pd.ExcelWriter(nom_excel_month, engine='openpyxl', mode='a') as writer:
            try:
                writer.book.remove(writer.book[actual_day])                  
            except KeyError:
                    pass
        with pd.ExcelWriter(nom_excel_month, engine='openpyxl', mode='a') as writer:
            try:
                writer.book.remove(writer.book["Data for WP"])
            except KeyError:
                pass

            df1.to_excel(writer, sheet_name=f"ALL", index=False)
            df2.to_excel(writer, sheet_name=actual_day, index=False)
            df3.to_excel(writer, sheet_name="Data for WP", index=False)

            print(f"Scrapping terminé !")

    else:
        print("Erreur", result.status_code)