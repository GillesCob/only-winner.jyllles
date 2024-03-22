from html2image import Html2Image
from datetime import datetime
from openpyxl import load_workbook
from flask import Flask, render_template



hti = Html2Image(size=(1000,1500))

#Données de départ
scrapping_month = "Avril"
actual_day = str(datetime.now().day)
Feuille_datas = str(actual_day)

#Chemins vers mes 2 fichiers excel
excel_data_file_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/EXCEL/DATAS.xlsx'
excel_BDD_INITIALE_file_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/BDD INITIALE.xlsx'



#Chargement des fichiers Excel
excel_month_workbook = load_workbook(excel_data_file_path)
excel_BDD_INITIALE_workbook = load_workbook(excel_BDD_INITIALE_file_path)

#Identification des feuilles utilisées par la suite
excel_sheet = excel_month_workbook[Feuille_datas]
nom_comp_of_sport_sheet = excel_BDD_INITIALE_workbook["COMP OF SPORT"] #Sert pour la donnée en haut de carte

#Mise en place des listes contenant les données récoltées dans les feuilles Excel
comp_of_sport_BDD_list = []
comp_of_sport_BDD_card_list = []
COMP_list = []
SPORT_list = []
PROMPT_list = []
PROMPT_list = []
NOM_NFT_list = []
COMP_OF_SPORT_list = []

#------------------Récupération des données dans l'Excel BDD INITIALE-----------------------------------#
#Je récupère toutes les valeurs présentes dans la colonne A de la feuille "COMP OF SPORT"
for cell in nom_comp_of_sport_sheet['A'][1:]:
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        comp_of_sport_BDD_list.append(cell.value)
        
#Je récupère toutes les valeurs présentes dans la colonne C (traduction de comp of sport pour le haut de la carte)
for cell in nom_comp_of_sport_sheet['C'][1:]:
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        comp_of_sport_BDD_card_list.append(cell.value)
        
#------------------Récupération des données dans l'Excel DU JOUR-----------------------------------#
for cell in excel_sheet['F'][1:]: 
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        COMP_list.append(cell.value)
        
for cell in excel_sheet['E'][1:]: 
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        SPORT_list.append(cell.value)
        
for comp, sport in zip(COMP_list, SPORT_list):
    COMP_OF_SPORT_list.append(f"{comp} of {sport}")
    
for cell in excel_sheet['L'][1:]: 
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        PROMPT_list.append(cell.value)

for cell in excel_sheet['M'][1:]: 
    # Vérifier si la cellule n'est pas vide et ajouter sa valeur à la liste
    if cell.value is not None:
        NOM_NFT_list.append(cell.value)
        



for index, competition in enumerate(COMP_OF_SPORT_list) :
    comp_of_sport = competition
    prompt = PROMPT_list[index]
    nft_name = NOM_NFT_list[index]
    #Chemin vers les images Midjourney
    image_midjourney_path = f'/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/IMAGES_MIDJOURNEY/{nft_name}.png'
    print(comp_of_sport)
    
    
    # Modèle de HTML avec CSS
    html_template = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Rectangles</title>
        <style>
            /* Ajoutez votre CSS ici */
            .container {{
                position: fixed;
                top: 0;
                left: 0;
                display: flex;
                justify-content: center;
                align-items: center;
                width: 100%;
            }}

            .black-rectangle {{
                width: 1000px;
                height: 1500px;
                background-color: black;
            }}

            .orange-rectangle {{
                position: absolute;
                width: 975px;
                height: 1475px;
                background: linear-gradient(to top right, #d8aa3f, #927021);
            }}

            .competition {{
                position: absolute;
                top: 37.5px;
                left: 50%;
                transform: translateX(-50%);
                color: rgb(0, 0, 0);
                font-size: 45px;
            }}

            .année {{
                position: absolute;
                top: 100px;
                left: 50%;
                transform: translateX(-50%);
                color: rgb(0, 0, 0);
                font-size: 50px;
            }}

            .fond-prompt {{
                position: absolute;
                width: 897px;
                height: 500px;
                bottom: 50px;
                background: hsla(0, 0%, 71%, 0.7);
                border-radius: 20px;
                border: 1px solid black;
            }}

            .prompt {{
                position: absolute;
                top: 1092px;
                left: 50%;
                transform: translateX(-50%);
                width: 850px;
                height: 343px;
                color: black;
                font-size: 40px;
                text-align: center;
                display: flex;
                justify-content: center;
                align-items: center;
                line-height: 50px;
            }}

            .image {{
                position: absolute;
                top: 178px;
                left: 50%;
                transform: translate(-50%);
                width: 900px;
                height: 900px;
                max-width: 100%;
                max-height: 100%;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="black-rectangle"></div>
            <div class="orange-rectangle"></div>
            <div class="competition">{comp_of_sport}</div>
            <div class="année">2024</div>
            <div class="fond-prompt"></div>
            <img src="{image_midjourney_path}" alt="Image" class="image">
            <div class="prompt">{prompt}</div>
        </div>
    </body>
    </html>
    """
    scrapping_month = "Avril"
    # Chemin vers le dossier de sauvegarde (fichier général + fichier du jour pour faciliter l'import des images sur WP)
    save_folder_paths = [f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/NFT_READY", f"/Users/gillescobigo/Documents/Gilles/Dev/Only Winners/DATAS/2024/{scrapping_month}/NFT_READY/IMPORT WP/{actual_day}"]
    
    for save_folder_path in save_folder_paths :
        # Définir le dossier de sortie
        hti.output_path = save_folder_path

        # Prendre une capture d'écran avec le HTML généré pour la compétition actuelle
        hti.screenshot(html_str=html_template, save_as=f'{competition}.png')